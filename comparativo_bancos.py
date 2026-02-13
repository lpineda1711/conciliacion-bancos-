import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import unicodedata

st.title("Conciliación Bancaria")

# Subir archivo Excel
uploaded_file = st.file_uploader("Sube tu archivo Excel (.xlsx)", type=["xlsx"])

if uploaded_file:
    xls = pd.ExcelFile(uploaded_file)

    # Verificar que existan las hojas
    if "Bancos" in xls.sheet_names and "Mayor" in xls.sheet_names:
        df_bancos = pd.read_excel(xls, sheet_name="Bancos")
        df_mayor = pd.read_excel(xls, sheet_name="Mayor")

        # 🔹 Normalizar columnas: minúsculas, sin espacios ni acentos
        def clean_columns(cols):
            return [
                unicodedata.normalize('NFKD', str(c))
                .encode('ascii', errors='ignore').decode('utf-8')
                .strip().lower().replace(' ', '') for c in cols
            ]

        df_bancos.columns = clean_columns(df_bancos.columns)
        df_mayor.columns = clean_columns(df_mayor.columns)

        # 🔹 Mostrar columnas detectadas (para depuración)
        st.write("Columnas detectadas en Bancos:", df_bancos.columns.tolist())
        st.write("Columnas detectadas en Mayor:", df_mayor.columns.tolist())

        # 🔹 Buscar columnas clave
        try:
            bancos_col = [c for c in df_bancos.columns if 'bancos' in c][0]  # columna C/D
            valor_col = [c for c in df_bancos.columns if 'valor' in c][0]
            debe_col = [c for c in df_mayor.columns if 'debe' in c][0]
            haber_col = [c for c in df_mayor.columns if 'haber' in c][0]
        except IndexError:
            st.error("No se pudieron encontrar las columnas clave. Revisa que tu Excel tenga 'Bancos', 'Valor', 'Debe' y 'Haber'.")
            st.stop()

        # 🔹 Crear resumen de valores que no coinciden
        resumen = []
        for index, row in df_bancos.iterrows():
            tipo = str(row[bancos_col]).strip().upper()
            valor = row[valor_col]

            if tipo == 'C':  # Debe
                if df_mayor[df_mayor[debe_col] == valor].empty:
                    resumen.append({'Fila': index + 2, 'Tipo': 'Debe', 'Valor': valor})
            elif tipo == 'D':  # Haber
                if df_mayor[df_mayor[haber_col] == valor].empty:
                    resumen.append({'Fila': index + 2, 'Tipo': 'Haber', 'Valor': valor})

        # 🔹 Guardar Excel con hojas originales y resumen
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_bancos.to_excel(writer, index=False, sheet_name='Bancos')
            df_mayor.to_excel(writer, index=False, sheet_name='Mayor')
            df_resumen = pd.DataFrame(resumen)
            df_resumen.to_excel(writer, index=False, sheet_name='Resumen')
            writer.save()
        output.seek(0)

        # 🔹 Abrir con openpyxl para sombrear valores del resumen
        wb = load_workbook(output)
        if 'Resumen' in wb.sheetnames:
            ws_resumen = wb['Resumen']
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            for row in range(2, ws_resumen.max_row + 1):
                ws_resumen[f'C{row}'].fill = yellow_fill  # columna Valor

        # 🔹 Guardar cambios finales y permitir descarga
        final_output = BytesIO()
        wb.save(final_output)
        final_output.seek(0)

        st.download_button(
            label="Descargar Excel Conciliado",
            data=final_output,
            file_name="Conciliacion.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    else:
        st.error("El archivo debe contener las hojas 'Bancos' y 'Mayor'.")




